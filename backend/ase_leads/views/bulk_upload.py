"""
Bulk Upload and Add Lead Views

POST /api/ase-leads/add-lead/          - Add a single lead (name, phone, location)
POST /api/ase-leads/bulk-upload/       - Bulk upload leads from Excel file
GET  /api/ase-leads/bulk-upload/template/  - Download Excel template
"""

import io
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status

from ase_leads.models import ASELead
from ase_leads.models.bre_data import BREResearchData
from ase_leads.permissions import ASEMarketingPermission


@api_view(['POST'])
@permission_classes([IsAuthenticated, ASEMarketingPermission])
def add_lead(request):
    """
    Add a single lead with minimal fields: name, phone, location.

    Request body:
      - name (required): Contact person name
      - phone (required): Phone number
      - location (optional): Location/city
      - notes (optional): Additional notes

    Creates an ASELead with status='new' and the user's company.
    """
    name = request.data.get('name', '').strip()
    phone = request.data.get('phone', '').strip()
    location = request.data.get('location', '').strip()
    notes = request.data.get('notes', '').strip()

    if not name:
        return Response({'error': 'name is required.'}, status=status.HTTP_400_BAD_REQUEST)
    if not phone:
        return Response({'error': 'phone is required.'}, status=status.HTTP_400_BAD_REQUEST)

    # Check for duplicate phone - show who created it
    company = request.user.company
    # For admin without company, try to get company from query params or default to ASE
    if not company:
        company_id = request.query_params.get('company') or request.data.get('company')
        if company_id:
            from accounts.models import Company
            try:
                company = Company.objects.get(id=company_id)
            except Company.DoesNotExist:
                return Response({'error': 'Company not found.'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'error': 'Company is required for admin users.'}, status=status.HTTP_400_BAD_REQUEST)
    existing = BREResearchData.objects.filter(phone_number=phone, company=company).select_related('created_by').first()
    if existing:
        creator_name = f"{existing.created_by.first_name} {existing.created_by.last_name}".strip() if existing.created_by else 'Unknown'
        return Response(
            {'error': f'The number {phone} is already created by "{creator_name}"'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Build notes field: location + notes combined
    full_notes = notes if notes else ''

    lead = BREResearchData.objects.create(
        name=name,
        phone_number=phone,
        location=location,
        notes=full_notes,
        company=company,
        created_by=request.user,
        assigned_to=request.user,  # Auto-assign to creator
    )

    return Response({
        'id': lead.id,
        'name': lead.name,
        'phone': lead.phone_number,
        'location': lead.location,
        'notes': lead.notes,
        'created_by': lead.created_by_name,
        'assigned_to': lead.assigned_to_name,
        'message': 'Lead added successfully',
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([IsAuthenticated, ASEMarketingPermission])
@parser_classes([MultiPartParser, FormParser])
def bulk_upload(request):
    """
    Bulk upload leads from an Excel (.xlsx) or CSV file.

    Expected columns: name, phone, location (or phone_number)
    
    Returns count of successfully created leads and any errors.
    """
    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)

    filename = file.name.lower()
    company = request.user.company

    try:
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            import openpyxl
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return Response({'error': 'File is empty.'}, status=status.HTTP_400_BAD_REQUEST)
            
            # First row is header
            headers = [str(h).lower().strip() if h else '' for h in rows[0]]
            data_rows = rows[1:]

        elif filename.endswith('.csv'):
            import csv
            content = file.read().decode('utf-8')
            reader = csv.reader(io.StringIO(content))
            rows = list(reader)
            if not rows:
                return Response({'error': 'File is empty.'}, status=status.HTTP_400_BAD_REQUEST)
            
            headers = [h.lower().strip() for h in rows[0]]
            data_rows = rows[1:]
        else:
            return Response(
                {'error': 'Unsupported file format. Please upload .xlsx, .xls, or .csv file.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Map column names
        name_col = None
        phone_col = None
        location_col = None

        for i, h in enumerate(headers):
            if h in ('name', 'contact_person', 'company_name', 'contact name'):
                name_col = i
            elif h in ('phone', 'phone_number', 'phone number', 'mobile', 'contact number'):
                phone_col = i
            elif h in ('location', 'city', 'address', 'area'):
                location_col = i

        if name_col is None:
            return Response(
                {'error': 'Could not find "name" column. Expected columns: name, phone, location'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if phone_col is None:
            return Response(
                {'error': 'Could not find "phone" column. Expected columns: name, phone, location'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Process rows
        created = 0
        errors = []
        duplicates = 0

        for row_num, row in enumerate(data_rows, start=2):
            try:
                # Skip completely empty rows
                if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                    continue

                name = str(row[name_col]).strip() if row[name_col] else ''
                phone = str(row[phone_col]).strip() if row[phone_col] else ''
                location = str(row[location_col]).strip() if location_col is not None and len(row) > location_col and row[location_col] else ''

                # Clean phone number (remove .0 from Excel numbers)
                if phone.endswith('.0'):
                    phone = phone[:-2]
                phone = phone.replace(' ', '').replace('-', '').replace('+91', '')

                # Skip rows where both name and phone are empty
                if not name and not phone:
                    continue

                if not name or not phone:
                    errors.append(f'Row {row_num}: Missing name or phone')
                    continue

                if len(phone) < 10:
                    errors.append(f'Row {row_num}: Invalid phone number "{phone}"')
                    continue

                # Check duplicate - show who created it
                existing = BREResearchData.objects.filter(phone_number=phone, company=company).select_related('created_by').first()
                if existing:
                    creator_name = f"{existing.created_by.first_name} {existing.created_by.last_name}".strip() if existing.created_by else 'Unknown'
                    duplicates += 1
                    errors.append(f'Row {row_num}: Phone {phone} already created by "{creator_name}"')
                    continue

                BREResearchData.objects.create(
                    name=name,
                    phone_number=phone,
                    location=location,
                    notes='',
                    company=company,
                    created_by=request.user,
                    assigned_to=request.user,  # Auto-assign to creator
                )
                created += 1

            except Exception as e:
                errors.append(f'Row {row_num}: {str(e)}')

        return Response({
            'message': f'Upload complete. {created} leads created.',
            'created': created,
            'duplicates': duplicates,
            'errors': errors[:20],  # Limit error messages
            'total_rows': len(data_rows),
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        return Response(
            {'error': f'Failed to process file: {str(e)}'},
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['GET'])
@permission_classes([])
def download_template(request):
    """
    Download an Excel template for bulk upload.
    Phone column is formatted as text to preserve leading zeros.
    """
    import openpyxl
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Leads'
    
    # Headers
    ws.append(['name', 'phone', 'location'])
    
    # Format phone column (B) as text for data rows only
    for row in range(2, 6):  # Only format the sample rows
        ws.cell(row=row, column=2).number_format = '@'
    
    # Sample data - write phone as string
    ws.cell(row=2, column=1, value='John Doe')
    ws.cell(row=2, column=2, value='9876543210')
    ws.cell(row=2, column=2).number_format = '@'
    ws.cell(row=2, column=3, value='Mumbai')
    
    ws.cell(row=3, column=1, value='Jane Smith')
    ws.cell(row=3, column=2, value='8765432109')
    ws.cell(row=3, column=2).number_format = '@'
    ws.cell(row=3, column=3, value='Delhi')
    
    ws.cell(row=4, column=1, value='Raj Kumar')
    ws.cell(row=4, column=2, value='7654321098')
    ws.cell(row=4, column=2).number_format = '@'
    ws.cell(row=4, column=3, value='Bangalore')
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="leads_upload_template.xlsx"'
    return response


# ══════════════════════════════════════════════════════════════════════════════
# BRE Research Data List/CRUD Endpoints
# ══════════════════════════════════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def bre_research_list(request):
    """
    List BRE research data with search, pagination.
    Admin sees all records; team members see their company's records.
    
    Query params:
      - search: search by name, phone_number, location
      - page: page number (default 1)
      - page_size: items per page (default 50)
    """
    from django.db.models import Q, Case, When, IntegerField
    from django.core.paginator import Paginator

    user = request.user
    # Admin sees all records, optionally filtered by company query param
    if user.role == 'admin':
        company_id = request.query_params.get('company', '').strip()
        if company_id:
            qs = BREResearchData.objects.filter(company_id=company_id)
        elif user.company:
            # Default to ASE Technologies company for the marketing panel
            from accounts.models import Company
            ase_company = Company.objects.filter(code__in=['ASE', 'ASE_TECH']).first()
            if ase_company:
                qs = BREResearchData.objects.filter(company=ase_company)
            else:
                qs = BREResearchData.objects.all()
        else:
            qs = BREResearchData.objects.all()
    elif user.role == 'manager':
        # Manager sees all records from their company
        company = user.company
        qs = BREResearchData.objects.filter(company=company)
    else:
        # Employees see only their own data (created_by = current user)
        company = user.company
        qs = BREResearchData.objects.filter(company=company, created_by=user)
    
    qs = qs.select_related('created_by', 'assigned_to').annotate(
        status_order=Case(
            When(status='new', then=0),
            default=1,
            output_field=IntegerField(),
        )
    ).order_by('status_order', '-created_at')

    # Search
    search = request.query_params.get('search', '').strip()
    if search:
        qs = qs.filter(
            Q(name__icontains=search) |
            Q(phone_number__icontains=search) |
            Q(location__icontains=search)
        )

    # Status filter
    status_filter = request.query_params.get('status', '').strip()
    if status_filter and status_filter != 'all':
        qs = qs.filter(status=status_filter)

    # Assigned to filter (filter by specific employee)
    assigned_to_filter = request.query_params.get('assigned_to', '').strip()
    if assigned_to_filter:
        qs = qs.filter(assigned_to_id=assigned_to_filter)

    # Created by filter (filter by which employee uploaded the data)
    created_by_filter = request.query_params.get('created_by', '').strip()
    if created_by_filter:
        qs = qs.filter(created_by_id=created_by_filter)

    # Date filter
    date_from = request.query_params.get('date_from', '').strip()
    date_to = request.query_params.get('date_to', '').strip()
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    # Pagination
    page_size = int(request.query_params.get('page_size', 50))
    page_number = int(request.query_params.get('page', 1))
    paginator = Paginator(qs, page_size)
    page = paginator.get_page(page_number)

    results = []
    for item in page.object_list:
        results.append({
            'id': item.id,
            'name': item.name,
            'phone_number': item.phone_number,
            'location': item.location,
            'notes': item.notes,
            'status': item.status,
            'created_by_name': item.created_by_name,
            'assigned_to_name': item.assigned_to_name,
            'created_by_id': item.created_by_id,
            'assigned_to_id': item.assigned_to_id,
            'created_at': item.created_at.isoformat(),
        })

    return Response({
        'results': results,
        'count': paginator.count,
        'page': page.number,
        'total_pages': paginator.num_pages,
    })


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def bre_research_update(request, pk):
    """Update a BRE research data record."""
    user = request.user
    if user.role == 'admin':
        try:
            item = BREResearchData.objects.get(pk=pk)
        except BREResearchData.DoesNotExist:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
    else:
        # Employees can only update their own data
        try:
            item = BREResearchData.objects.get(pk=pk, company=user.company, created_by=user)
        except BREResearchData.DoesNotExist:
            return Response({'error': 'Not found or you do not have permission.'}, status=status.HTTP_404_NOT_FOUND)

    # Update fields
    if 'name' in request.data:
        item.name = request.data['name']
    if 'phone_number' in request.data:
        item.phone_number = request.data['phone_number']
    if 'location' in request.data:
        item.location = request.data['location']
    if 'notes' in request.data:
        item.notes = request.data['notes']
    if 'assigned_to' in request.data:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        try:
            item.assigned_to = User.objects.get(pk=request.data['assigned_to'])
            item.status = 'assigned'
        except User.DoesNotExist:
            pass

    item.save()
    return Response({
        'id': item.id,
        'name': item.name,
        'phone_number': item.phone_number,
        'location': item.location,
        'notes': item.notes,
        'created_by_name': item.created_by_name,
        'assigned_to_name': item.assigned_to_name,
    })


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def bre_research_delete(request, pk):
    """Delete a BRE research data record."""
    user = request.user
    if user.role == 'admin':
        try:
            item = BREResearchData.objects.get(pk=pk)
        except BREResearchData.DoesNotExist:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
    else:
        # Employees can only delete their own data
        try:
            item = BREResearchData.objects.get(pk=pk, company=user.company, created_by=user)
        except BREResearchData.DoesNotExist:
            return Response({'error': 'Not found or you do not have permission.'}, status=status.HTTP_404_NOT_FOUND)

    item.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bre_research_bulk_delete(request):
    """
    Bulk delete multiple BRE research records.
    """
    from django.db.models import Q

    select_all = request.data.get('select_all', False)
    user = request.user
    
    # Admin operates on ASE records
    if user.role == 'admin':
        from accounts.models import Company
        ase_company = Company.objects.filter(code__in=['ASE', 'ASE_TECH']).first()
        if ase_company:
            base_qs = BREResearchData.objects.filter(company=ase_company)
        else:
            base_qs = BREResearchData.objects.all()
    else:
        # Employees can only delete their own data
        base_qs = BREResearchData.objects.filter(company=user.company, created_by=user)

    if select_all:
        # Apply same filters as list endpoint
        qs = base_qs.all()
        search = request.data.get('search', '').strip()
        if search:
            qs = qs.filter(Q(name__icontains=search) | Q(phone_number__icontains=search) | Q(location__icontains=search))
        status_filter = request.data.get('status', '').strip()
        if status_filter and status_filter != 'all':
            qs = qs.filter(status=status_filter)
        assigned_to_filter = request.data.get('assigned_to', '').strip() if request.data.get('assigned_to') else ''
        if assigned_to_filter:
            qs = qs.filter(assigned_to_id=assigned_to_filter)
        date_from = request.data.get('date_from', '').strip() if request.data.get('date_from') else ''
        date_to = request.data.get('date_to', '').strip() if request.data.get('date_to') else ''
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        deleted_count, _ = qs.delete()
    else:
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'error': 'ids is required (list of record IDs).'}, status=status.HTTP_400_BAD_REQUEST)
        deleted_count, _ = base_qs.filter(id__in=ids).delete()

    return Response({
        'message': f'{deleted_count} records deleted successfully.',
        'deleted': deleted_count,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bre_research_bulk_assign(request):
    """
    Bulk assign multiple BRE research records to a BOE team member.
    """
    from django.contrib.auth import get_user_model
    from django.db.models import Q
    User = get_user_model()

    assigned_to_id = request.data.get('assigned_to')
    select_all = request.data.get('select_all', False)

    if not assigned_to_id:
        return Response({'error': 'assigned_to is required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        assigned_user = User.objects.get(pk=assigned_to_id)
    except User.DoesNotExist:
        return Response({'error': 'Assigned user not found.'}, status=status.HTTP_400_BAD_REQUEST)

    user = request.user
    # Admin operates on ASE records
    if user.role == 'admin':
        from accounts.models import Company
        ase_company = Company.objects.filter(code__in=['ASE', 'ASE_TECH']).first()
        if ase_company:
            base_qs = BREResearchData.objects.filter(company=ase_company)
        else:
            base_qs = BREResearchData.objects.all()
    else:
        # Employees can only assign their own data
        base_qs = BREResearchData.objects.filter(company=user.company, created_by=user)

    if select_all:
        qs = base_qs.all()
        search = request.data.get('search', '').strip()
        if search:
            qs = qs.filter(Q(name__icontains=search) | Q(phone_number__icontains=search) | Q(location__icontains=search))
        status_filter = request.data.get('status', '').strip()
        if status_filter and status_filter != 'all':
            qs = qs.filter(status=status_filter)
        assigned_to_filter = str(request.data.get('assigned_to_filter', '')).strip()
        if assigned_to_filter:
            qs = qs.filter(assigned_to_id=assigned_to_filter)
        date_from = str(request.data.get('date_from', '')).strip()
        date_to = str(request.data.get('date_to', '')).strip()
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        # Limit: assign only a specific number of records
        limit = request.data.get('limit')
        if limit:
            limit = int(limit)
            ids_to_update = list(qs.values_list('id', flat=True)[:limit])
            updated = BREResearchData.objects.filter(id__in=ids_to_update).update(assigned_to=assigned_user, status='assigned')
        else:
            updated = qs.update(assigned_to=assigned_user, status='assigned')
    else:
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'error': 'ids is required (list of record IDs).'}, status=status.HTTP_400_BAD_REQUEST)
        updated = base_qs.filter(id__in=ids).update(assigned_to=assigned_user, status='assigned')

    return Response({
        'message': f'{updated} records assigned to {assigned_user.first_name} {assigned_user.last_name}'.strip(),
        'updated': updated,
    })


# ══════════════════════════════════════════════════════════════════════════════
# BOE Assigned Data Endpoint
# ══════════════════════════════════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def boe_assigned_list(request):
    """
    List BRE research data assigned to the current BOE user.
    Shows data that has been assigned to this user by BRE employees.
    
    Query params:
      - search: search by name, phone_number, location
      - page: page number (default 1)
      - page_size: items per page (default 50)
      - date_from: filter by assignment date
      - date_to: filter by assignment date
    """
    from django.db.models import Q
    from django.core.paginator import Paginator
    from django.utils import timezone
    from datetime import timedelta

    user = request.user
    company = user.company
    
    # Admin sees all ASE assigned data; BOE sees only their own
    if user.role == 'admin':
        from accounts.models import Company
        ase_company = Company.objects.filter(code__in=['ASE', 'ASE_TECH']).first()
        if ase_company:
            qs = BREResearchData.objects.filter(
                company=ase_company,
                status__in=['assigned', 'converted'],
            ).select_related('created_by', 'assigned_to', 'assigned_to_cre')
        else:
            qs = BREResearchData.objects.filter(
                status__in=['assigned', 'converted'],
            ).select_related('created_by', 'assigned_to', 'assigned_to_cre')
    elif user.role == 'manager':
        qs = BREResearchData.objects.filter(
            company=company,
            status__in=['assigned', 'converted'],
        ).select_related('created_by', 'assigned_to', 'assigned_to_cre')
    else:
        # BOE employee sees only their assigned data
        qs = BREResearchData.objects.filter(
            company=company,
            assigned_to=user,
            status__in=['assigned', 'converted'],
        ).select_related('created_by', 'assigned_to', 'assigned_to_cre')

    # Order: assigned (active) records at top, converted at bottom.
    # Records with call_status='interested' (shown as "Converted" in UI) also go to bottom.
    # Within each group, newest first.
    from django.db.models import Case, When, Value, IntegerField
    qs = qs.annotate(
        status_order=Case(
            When(status='converted', then=Value(1)),
            When(call_status='interested', then=Value(1)),
            default=Value(0),
            output_field=IntegerField(),
        )
    ).order_by('status_order', '-created_at')

    # Search
    search = request.query_params.get('search', '').strip()
    if search:
        qs = qs.filter(
            Q(name__icontains=search) |
            Q(phone_number__icontains=search) |
            Q(location__icontains=search)
        )

    # Date filter
    date_from = request.query_params.get('date_from', '').strip()
    date_to = request.query_params.get('date_to', '').strip()
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    # Call status filter
    call_status_filter = request.query_params.get('call_status', '').strip()
    if call_status_filter:
        qs = qs.filter(call_status=call_status_filter)

    # Calculate stats
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)

    if user.role == 'admin':
        from accounts.models import Company as CompanyModel
        ase_co = CompanyModel.objects.filter(code__in=['ASE', 'ASE_TECH']).first()
        if ase_co:
            all_assigned = BREResearchData.objects.filter(company=ase_co, status='assigned')
        else:
            all_assigned = BREResearchData.objects.filter(status='assigned')
    elif user.role == 'manager':
        all_assigned = BREResearchData.objects.filter(company=company, status='assigned')
    else:
        all_assigned = BREResearchData.objects.filter(company=company, assigned_to=user, status='assigned')
    stats = {
        'total_assigned': all_assigned.count(),
        'today_assigned': all_assigned.filter(created_at__date=today).count(),
        'this_week_assigned': all_assigned.filter(created_at__date__gte=week_start).count(),
        'this_month_assigned': all_assigned.filter(created_at__date__gte=month_start).count(),
    }

    # Pagination
    page_size = int(request.query_params.get('page_size', 50))
    page_number = int(request.query_params.get('page', 1))
    paginator = Paginator(qs, page_size)
    page = paginator.get_page(page_number)

    results = []
    for item in page.object_list:
        cre_name = None
        if item.assigned_to_cre:
            cre_name = f"{item.assigned_to_cre.first_name} {item.assigned_to_cre.last_name}".strip() or item.assigned_to_cre.username
        results.append({
            'id': item.id,
            'name': item.name,
            'phone_number': item.phone_number,
            'location': item.location,
            'notes': item.notes,
            'status': item.status,
            'call_status': item.call_status,
            'call_notes': item.call_notes,
            'created_by_name': item.created_by_name,
            'assigned_to_name': item.assigned_to_name,
            'assigned_to_cre_name': cre_name,
            'created_at': item.created_at.isoformat(),
        })

    return Response({
        'results': results,
        'count': paginator.count,
        'page': page.number,
        'total_pages': paginator.num_pages,
        'stats': stats,
    })


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def boe_update_call_status(request, pk):
    """
    BOE employee updates call status for an assigned record.
    Can set call_status and call_notes.
    """
    user = request.user
    try:
        item = BREResearchData.objects.get(pk=pk, assigned_to=user)
    except BREResearchData.DoesNotExist:
        return Response({'error': 'Record not found or not assigned to you.'}, status=status.HTTP_404_NOT_FOUND)

    call_status = request.data.get('call_status')
    call_notes = request.data.get('call_notes')

    if call_status:
        valid_statuses = ['pending', 'no_answer', 'callback', 'not_interested', 'interested']
        if call_status not in valid_statuses:
            return Response({'error': f'Invalid call_status. Must be one of: {valid_statuses}'}, status=status.HTTP_400_BAD_REQUEST)
        item.call_status = call_status

        # When marked as "interested", auto-create a lead in the separate leads table
        if call_status == 'interested':
            from ase_leads.models.boe_lead import BOELead
            # Only create if not already exists
            if not BOELead.objects.filter(source_research=item, created_by=user).exists():
                BOELead.objects.create(
                    name=item.name,
                    phone_number=item.phone_number,
                    location=item.location,
                    notes=item.notes,
                    call_notes=item.call_notes or '',
                    source_research=item,
                    created_by=user,
                    company=user.company,
                )

    if call_notes is not None:
        item.call_notes = call_notes

    item.save()
    return Response({
        'id': item.id,
        'call_status': item.call_status,
        'call_notes': item.call_notes,
        'message': 'Call status updated successfully.',
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def boe_convert_to_lead(request, pk):
    """
    BOE employee converts a positive call to a lead and assigns to CRE.
    Required: assigned_to_cre (CRE user ID)
    """
    from django.contrib.auth import get_user_model
    User = get_user_model()

    user = request.user
    try:
        item = BREResearchData.objects.get(pk=pk, assigned_to=user)
    except BREResearchData.DoesNotExist:
        return Response({'error': 'Record not found or not assigned to you.'}, status=status.HTTP_404_NOT_FOUND)

    cre_user_id = request.data.get('assigned_to_cre')
    if not cre_user_id:
        return Response({'error': 'assigned_to_cre is required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        cre_user = User.objects.get(pk=cre_user_id)
    except User.DoesNotExist:
        return Response({'error': 'CRE user not found.'}, status=status.HTTP_400_BAD_REQUEST)

    # Update the record
    item.status = 'converted'
    item.call_status = 'interested'
    item.assigned_to_cre = cre_user
    if request.data.get('call_notes'):
        item.call_notes = request.data['call_notes']
    item.save()

    cre_name = f"{cre_user.first_name} {cre_user.last_name}".strip() or cre_user.username
    return Response({
        'id': item.id,
        'status': item.status,
        'call_status': item.call_status,
        'assigned_to_cre_name': cre_name,
        'message': f'Lead converted and assigned to {cre_name} successfully.',
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def cre_users_list(request):
    """List all CRE team members and managers/team_leads/admins for assignment dropdown."""
    from django.contrib.auth import get_user_model
    from django.db.models import Q
    from teams.models import Team
    User = get_user_model()

    # Get CRE team
    cre_team = Team.objects.filter(marketing_category='cre', is_active=True).first()

    # Get all CRE team members (any role) + managers/team_leads/admins
    filter_q = Q(role__in=['manager', 'team_lead', 'admin'])
    if cre_team:
        filter_q = filter_q | Q(team=cre_team)

    users = User.objects.filter(
        is_active=True
    ).filter(filter_q).distinct().values('id', 'first_name', 'last_name', 'username', 'role')

    result = []
    for u in users:
        name = f"{u['first_name']} {u['last_name']}".strip() or u['username']
        role_label = u['role'].replace('_', ' ').title() if u['role'] != 'employee' else ''
        display_name = f"{name} ({role_label})" if role_label else name
        result.append({'id': u['id'], 'name': display_name})
    return Response(result)


# ══════════════════════════════════════════════════════════════════════════════
# BOE CRUD Operations
# ══════════════════════════════════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def boe_add_data(request):
    """BOE employee adds their own research data."""
    user = request.user
    name = request.data.get('name', '').strip()
    phone = request.data.get('phone_number', '').strip()
    location = request.data.get('location', '').strip()
    notes = request.data.get('notes', '').strip()

    if not name or not phone:
        return Response({'error': 'Name and phone number are required.'}, status=status.HTTP_400_BAD_REQUEST)

    company = user.company
    # Check duplicate
    existing = BREResearchData.objects.filter(phone_number=phone, company=company).first()
    if existing:
        creator_name = f"{existing.created_by.first_name} {existing.created_by.last_name}".strip() if existing.created_by else 'Unknown'
        return Response({'error': f'Phone number {phone} already exists (created by {creator_name})'}, status=status.HTTP_400_BAD_REQUEST)

    item = BREResearchData.objects.create(
        name=name,
        phone_number=phone,
        location=location,
        notes=notes,
        status='assigned',
        created_by=user,
        assigned_to=user,
        company=company,
    )
    return Response({
        'id': item.id,
        'name': item.name,
        'phone_number': item.phone_number,
        'message': 'Data added successfully.',
    }, status=status.HTTP_201_CREATED)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def boe_edit_data(request, pk):
    """BOE employee edits their own data or data assigned to them."""
    user = request.user
    try:
        item = BREResearchData.objects.get(pk=pk, assigned_to=user)
    except BREResearchData.DoesNotExist:
        return Response({'error': 'Record not found or not assigned to you.'}, status=status.HTTP_404_NOT_FOUND)

    if 'name' in request.data:
        item.name = request.data['name']
    if 'phone_number' in request.data:
        # Check duplicate
        new_phone = request.data['phone_number'].strip()
        if new_phone != item.phone_number:
            existing = BREResearchData.objects.filter(phone_number=new_phone, company=user.company).exclude(pk=pk).first()
            if existing:
                return Response({'error': f'Phone number {new_phone} already exists.'}, status=status.HTTP_400_BAD_REQUEST)
        item.phone_number = new_phone
    if 'location' in request.data:
        item.location = request.data['location']
    if 'notes' in request.data:
        item.notes = request.data['notes']

    item.save()
    return Response({
        'id': item.id,
        'name': item.name,
        'phone_number': item.phone_number,
        'location': item.location,
        'notes': item.notes,
        'message': 'Data updated successfully.',
    })


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def boe_delete_data(request, pk):
    """BOE employee deletes data they created (not data assigned from BRE)."""
    user = request.user
    try:
        # BOE can only delete data they created themselves
        item = BREResearchData.objects.get(pk=pk, assigned_to=user, created_by=user)
    except BREResearchData.DoesNotExist:
        return Response({'error': 'Record not found or you cannot delete data assigned from BRE.'}, status=status.HTTP_404_NOT_FOUND)

    item.delete()
    return Response({'message': 'Record deleted successfully.'}, status=status.HTTP_204_NO_CONTENT)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def boe_bulk_delete_data(request):
    """BOE employee bulk deletes data they created (not data assigned from BRE)."""
    user = request.user
    ids = request.data.get('ids', [])
    if not ids:
        return Response({'error': 'No IDs provided.'}, status=status.HTTP_400_BAD_REQUEST)

    # BOE can only delete data they created themselves
    items = BREResearchData.objects.filter(pk__in=ids, assigned_to=user, created_by=user)
    deleted_count = items.count()
    items.delete()
    return Response({'message': f'{deleted_count} record(s) deleted successfully.', 'deleted_count': deleted_count}, status=status.HTTP_200_OK)


# ══════════════════════════════════════════════════════════════════════════════
# BOE Leads Table CRUD (separate table)
# ══════════════════════════════════════════════════════════════════════════════

from ase_leads.models.boe_lead import BOELead


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def boe_leads_list(request):
    """List leads from the separate leads table. Admin sees all, BOE sees own."""
    from django.core.paginator import Paginator
    from django.db.models import Q

    user = request.user
    
    # Admin sees all ASE leads; BOE sees only their own
    if user.role == 'admin':
        from accounts.models import Company
        ase_company = Company.objects.filter(code__in=['ASE', 'ASE_TECH']).first()
        if ase_company:
            qs = BOELead.objects.filter(company=ase_company).select_related('assigned_to_cre', 'source_research', 'created_by')
        else:
            qs = BOELead.objects.all().select_related('assigned_to_cre', 'source_research', 'created_by')
    elif user.role == 'manager':
        qs = BOELead.objects.filter(company=user.company).select_related('assigned_to_cre', 'source_research', 'created_by')
    else:
        qs = BOELead.objects.filter(created_by=user).select_related('assigned_to_cre', 'source_research', 'created_by')

    # Search
    search = request.query_params.get('search', '').strip()
    if search:
        qs = qs.filter(Q(name__icontains=search) | Q(phone_number__icontains=search) | Q(location__icontains=search))

    # Status filter
    status_filter = request.query_params.get('status', '').strip()
    if status_filter and status_filter != 'all':
        qs = qs.filter(status=status_filter)

    # Date filter
    date_from = request.query_params.get('date_from', '').strip()
    date_to = request.query_params.get('date_to', '').strip()
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    # Created by / assigned to filter (admin/manager)
    created_by_filter = request.query_params.get('created_by', '').strip()
    if created_by_filter and user.role in ('admin', 'manager'):
        from django.db.models import Q
        qs = qs.filter(Q(created_by_id=created_by_filter) | Q(assigned_to_cre_id=created_by_filter))

    # CRE assigned filter
    assigned_cre_filter = request.query_params.get('assigned_cre', '').strip()
    if assigned_cre_filter:
        if assigned_cre_filter == 'unassigned':
            qs = qs.filter(assigned_to_cre__isnull=True)
        else:
            qs = qs.filter(assigned_to_cre_id=assigned_cre_filter)

    qs = qs.order_by('-created_at')

    # Pagination
    page_size = int(request.query_params.get('page_size', 50))
    page_number = int(request.query_params.get('page', 1))
    paginator = Paginator(qs, page_size)
    page = paginator.get_page(page_number)

    results = []
    for item in page.object_list:
        results.append({
            'id': item.id,
            'name': item.name,
            'phone_number': item.phone_number,
            'location': item.location,
            'notes': item.notes,
            'call_notes': item.call_notes,
            'status': item.status,
            'created_by_name': item.created_by_name,
            'assigned_to_cre_name': item.assigned_to_cre_name,
            'task_created': item.task_created,
            'created_at': item.created_at.isoformat(),
        })

    # Stats - based on filtered queryset (before pagination)
    stats = {
        'total': paginator.count,
        'pending_cre': qs.filter(status='interested').count(),
        'assigned_cre': qs.exclude(status='interested').count(),
    }

    return Response({
        'results': results,
        'count': paginator.count,
        'page': page.number,
        'total_pages': paginator.num_pages,
        'stats': stats,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def boe_leads_create(request):
    """BOE or Admin creates a lead directly."""
    user = request.user
    name = request.data.get('name', '').strip()
    phone = request.data.get('phone_number', '').strip()
    location = request.data.get('location', '').strip()
    notes = request.data.get('notes', '').strip()
    call_notes = request.data.get('call_notes', '').strip()
    source_id = request.data.get('source_research_id')

    if not name or not phone:
        return Response({'error': 'Name and phone are required.'}, status=status.HTTP_400_BAD_REQUEST)

    source = None
    if source_id:
        try:
            source = BREResearchData.objects.get(pk=source_id)
        except BREResearchData.DoesNotExist:
            pass

    # For admin without company, use ASE Technologies (ID 2) as default
    from accounts.models import Company
    company = user.company
    if not company:
        company = Company.objects.filter(code='ASE').first()

    lead = BOELead.objects.create(
        name=name,
        phone_number=phone,
        location=location,
        notes=notes,
        call_notes=call_notes,
        source_research=source,
        created_by=user,
        company=company,
    )
    return Response({
        'id': lead.id,
        'name': lead.name,
        'message': 'Lead created successfully.',
    }, status=status.HTTP_201_CREATED)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def boe_leads_update(request, pk):
    """BOE or Admin updates a lead."""
    user = request.user
    try:
        if user.role == 'admin':
            lead = BOELead.objects.get(pk=pk)
        else:
            lead = BOELead.objects.get(pk=pk, created_by=user)
    except BOELead.DoesNotExist:
        return Response({'error': 'Lead not found.'}, status=status.HTTP_404_NOT_FOUND)

    if 'name' in request.data:
        lead.name = request.data['name']
    if 'phone_number' in request.data:
        lead.phone_number = request.data['phone_number']
    if 'location' in request.data:
        lead.location = request.data['location']
    if 'notes' in request.data:
        lead.notes = request.data['notes']
    if 'call_notes' in request.data:
        lead.call_notes = request.data['call_notes']
    if 'status' in request.data:
        lead.status = request.data['status']

    lead.save()
    return Response({'id': lead.id, 'message': 'Lead updated.'})


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def boe_leads_delete(request, pk):
    """BOE or Admin deletes a lead."""
    user = request.user
    try:
        if user.role == 'admin':
            lead = BOELead.objects.get(pk=pk)
        else:
            lead = BOELead.objects.get(pk=pk, created_by=user)
    except BOELead.DoesNotExist:
        return Response({'error': 'Lead not found.'}, status=status.HTTP_404_NOT_FOUND)

    lead.delete()
    return Response({'message': 'Lead deleted.'}, status=status.HTTP_204_NO_CONTENT)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def boe_leads_assign_cre(request, pk):
    """BOE or Admin assigns a lead to CRE."""
    from django.contrib.auth import get_user_model
    from teams.models import Team
    User = get_user_model()

    user = request.user
    try:
        if user.role == 'admin':
            lead = BOELead.objects.get(pk=pk)
        else:
            lead = BOELead.objects.get(pk=pk, created_by=user)
    except BOELead.DoesNotExist:
        return Response({'error': 'Lead not found.'}, status=status.HTTP_404_NOT_FOUND)

    cre_user_id = request.data.get('assigned_to_cre')
    if not cre_user_id:
        return Response({'error': 'assigned_to_cre is required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        cre_user = User.objects.get(pk=cre_user_id)
    except User.DoesNotExist:
        return Response({'error': 'CRE user not found.'}, status=status.HTTP_400_BAD_REQUEST)

    # Validate that the user is a CRE team member or has manager/team_lead/admin role
    cre_team = Team.objects.filter(marketing_category='cre', is_active=True).first()
    allowed_roles = ('manager', 'team_lead', 'admin')
    is_cre_member = cre_team and cre_user.team_id == cre_team.id
    is_allowed_role = cre_user.role in allowed_roles

    if not is_cre_member and not is_allowed_role:
        return Response({'error': 'Selected user must be a CRE team member or have manager/team_lead/admin role.'}, status=status.HTTP_400_BAD_REQUEST)

    lead.assigned_to_cre = cre_user
    lead.status = 'assigned_cre'
    if request.data.get('call_notes'):
        lead.call_notes = request.data['call_notes']
    lead.save()

    cre_name = f"{cre_user.first_name} {cre_user.last_name}".strip() or cre_user.username
    return Response({'id': lead.id, 'message': f'Lead assigned to {cre_name}.'})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def boe_leads_mark_task_created(request, pk):
    """Mark a BOE lead as having a task created for it."""
    user = request.user
    try:
        if user.role in ('admin', 'manager', 'team_lead'):
            lead = BOELead.objects.get(pk=pk)
        else:
            lead = BOELead.objects.get(pk=pk, created_by=user)
    except BOELead.DoesNotExist:
        return Response({'error': 'Lead not found.'}, status=status.HTTP_404_NOT_FOUND)

    if lead.task_created:
        return Response({'error': 'Task already created for this lead.'}, status=status.HTTP_400_BAD_REQUEST)

    lead.task_created = True
    lead.save(update_fields=['task_created'])
    return Response({'id': lead.id, 'task_created': True, 'message': 'Lead marked as task created.'})# ══════════════════════════════════════════════════════════════════════════════
# BOE Leads Export / Import
# ══════════════════════════════════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def boe_leads_export(request):
    """Export BOE leads to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    user = request.user
    if user.role == 'admin':
        from accounts.models import Company
        ase_company = Company.objects.filter(code__in=['ASE', 'ASE_TECH']).first()
        qs = BOELead.objects.filter(company=ase_company) if ase_company else BOELead.objects.all()
    else:
        qs = BOELead.objects.filter(created_by=user)

    qs = qs.select_related('assigned_to_cre', 'created_by').order_by('-created_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'BOE Leads'

    # Header style
    header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    headers = ['Name', 'Phone Number', 'Location', 'Notes', 'Call Notes', 'Status', 'Created By', 'Assigned to CRE', 'Date']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 20

    # Phone column as text
    from openpyxl.styles import numbers
    for row_idx, lead in enumerate(qs, 2):
        ws.cell(row=row_idx, column=1, value=lead.name)
        phone_cell = ws.cell(row=row_idx, column=2, value=str(lead.phone_number))
        phone_cell.number_format = '@'
        ws.cell(row=row_idx, column=3, value=lead.location)
        ws.cell(row=row_idx, column=4, value=lead.notes)
        ws.cell(row=row_idx, column=5, value=lead.call_notes)
        ws.cell(row=row_idx, column=6, value=lead.status)
        ws.cell(row=row_idx, column=7, value=lead.created_by_name or '')
        ws.cell(row=row_idx, column=8, value=lead.assigned_to_cre_name or '')
        ws.cell(row=row_idx, column=9, value=lead.created_at.strftime('%Y-%m-%d'))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="boe_leads.xlsx"'
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def boe_leads_template(request):
    """Download Excel template for bulk import."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'BOE Leads Template'

    header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    headers = ['Name *', 'Phone Number *', 'Location', 'Notes', 'Call Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 22

    # Sample row
    ws.cell(row=2, column=1, value='John Doe')
    phone_cell = ws.cell(row=2, column=2, value='9876543210')
    phone_cell.number_format = '@'
    ws.cell(row=2, column=3, value='Hyderabad')
    ws.cell(row=2, column=4, value='Interested in property')
    ws.cell(row=2, column=5, value='Called twice')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="boe_leads_template.xlsx"'
    wb.save(response)
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def boe_leads_import(request):
    """Bulk import BOE leads from Excel."""
    import openpyxl
    from io import BytesIO

    user = request.user
    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

    from accounts.models import Company
    company = user.company
    if not company:
        company = Company.objects.filter(code='ASE').first()

    try:
        wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
        ws = wb.active

        created = 0
        skipped = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]:
                continue

            name = str(row[0]).strip() if row[0] else ''
            phone = str(row[1]).strip() if row[1] else ''
            location = str(row[2]).strip() if row[2] else ''
            notes = str(row[3]).strip() if row[3] else ''
            call_notes = str(row[4]).strip() if row[4] else ''

            if not name or not phone:
                errors.append(f'Row {row_idx}: Name and phone are required')
                skipped += 1
                continue

            # Clean phone number
            phone = phone.replace('.0', '').strip()

            BOELead.objects.create(
                name=name,
                phone_number=phone,
                location=location,
                notes=notes,
                call_notes=call_notes,
                created_by=user,
                company=company,
            )
            created += 1

        return Response({
            'message': f'{created} leads imported successfully.',
            'created': created,
            'skipped': skipped,
            'errors': errors[:10],
        })

    except Exception as e:
        return Response({'error': f'Failed to process file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


# ══════════════════════════════════════════════════════════════════════════════
# BRE Dashboard Stats Endpoint
# ══════════════════════════════════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def bre_dashboard_stats(request):
    """
    Return BRE dashboard statistics: total, new, assigned, today/week/month counts.
    Admin users see all records; team members see their company's records.
    """
    from django.utils import timezone
    from datetime import timedelta

    user = request.user
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)

    # Admin sees all ASE records
    if user.role == 'admin':
        from accounts.models import Company
        ase_company = Company.objects.filter(code__in=['ASE', 'ASE_TECH']).first()
        if ase_company:
            all_records = BREResearchData.objects.filter(company=ase_company)
        else:
            all_records = BREResearchData.objects.all()
    elif user.role == 'manager':
        company = user.company
        all_records = BREResearchData.objects.filter(company=company)
    else:
        # Employees see only their own data stats
        company = user.company
        all_records = BREResearchData.objects.filter(company=company, created_by=user)

    total = all_records.count()
    new_count = all_records.filter(status='new').count()
    assigned_count = all_records.filter(status='assigned').count()
    today_added = all_records.filter(created_at__date=today).count()
    this_week_added = all_records.filter(created_at__date__gte=week_start).count()
    this_week_assigned = all_records.filter(status='assigned', created_at__date__gte=week_start).count()
    this_month_added = all_records.filter(created_at__date__gte=month_start).count()
    this_month_assigned = all_records.filter(status='assigned', created_at__date__gte=month_start).count()

    return Response({
        'total': total,
        'new_count': new_count,
        'assigned_count': assigned_count,
        'today_added': today_added,
        'this_week_added': this_week_added,
        'this_week_assigned': this_week_assigned,
        'this_month_added': this_month_added,
        'this_month_assigned': this_month_assigned,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def bre_users_list(request):
    """
    List all BRE team members (creators) for filter dropdowns.
    Returns users who have created BRE research data.
    """
    from django.contrib.auth import get_user_model
    User = get_user_model()

    # Get distinct creator IDs from BRE research data
    creator_ids = BREResearchData.objects.values_list('created_by_id', flat=True).distinct()
    users = User.objects.filter(id__in=creator_ids, is_active=True).values(
        'id', 'first_name', 'last_name', 'email', 'username'
    )

    return Response(list(users))


# ══════════════════════════════════════════════════════════════════════════════
# CRE Leads Endpoint
# ══════════════════════════════════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def cre_leads_list(request):
    """
    List leads assigned to the current CRE user.
    Admin sees all CRE-assigned leads; CRE sees only their own.
    
    Query params:
      - search: search by name, phone_number, location
      - status: filter by lead status
      - date_from / date_to: filter by date
      - page / page_size: pagination
    """
    from django.db.models import Q
    from django.core.paginator import Paginator
    from django.utils import timezone
    from datetime import timedelta

    user = request.user
    company = user.company

    # Admin sees all ASE CRE-assigned leads; manager/team_lead see their company; employee sees own
    if user.role == 'admin':
        from accounts.models import Company
        ase_company = Company.objects.filter(code__in=['ASE', 'ASE_TECH']).first()
        if ase_company:
            qs = BOELead.objects.filter(company=ase_company, assigned_to_cre__isnull=False)
        else:
            qs = BOELead.objects.filter(assigned_to_cre__isnull=False)
    elif user.role in ('manager', 'team_lead'):
        if company:
            qs = BOELead.objects.filter(company=company, assigned_to_cre__isnull=False)
        else:
            qs = BOELead.objects.filter(assigned_to_cre__isnull=False)
    else:
        qs = BOELead.objects.filter(assigned_to_cre=user, company=company)

    qs = qs.select_related('assigned_to_cre', 'created_by').order_by('-created_at')

    # Search
    search = request.query_params.get('search', '').strip()
    if search:
        qs = qs.filter(
            Q(name__icontains=search) |
            Q(phone_number__icontains=search) |
            Q(location__icontains=search)
        )

    # Status filter
    status_filter = request.query_params.get('status', '').strip()
    if status_filter and status_filter != 'all':
        qs = qs.filter(status=status_filter)

    # Date filter
    date_from = request.query_params.get('date_from', '').strip()
    date_to = request.query_params.get('date_to', '').strip()
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    # Stats
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)

    if user.role == 'admin':
        from accounts.models import Company as CompanyMdl
        ase_co2 = CompanyMdl.objects.filter(code__in=['ASE', 'ASE_TECH']).first()
        all_cre = BOELead.objects.filter(company=ase_co2, assigned_to_cre__isnull=False) if ase_co2 else BOELead.objects.filter(assigned_to_cre__isnull=False)
    elif user.role == 'team_lead':
        all_cre = BOELead.objects.filter(company=company, assigned_to_cre__isnull=False) if company else BOELead.objects.filter(assigned_to_cre__isnull=False)
    else:
        all_cre = BOELead.objects.filter(assigned_to_cre=user, company=company)

    stats = {
        'total': all_cre.count(),
        'cold': all_cre.filter(status='cold').count(),
        'warm': all_cre.filter(status='warm').count(),
        'hot': all_cre.filter(status='hot').count(),
        'completed': all_cre.filter(status='completed').count(),
        'rejected': all_cre.filter(status='rejected').count(),
        'today_assigned': all_cre.filter(created_at__date=today).count(),
        'this_week': all_cre.filter(created_at__date__gte=week_start).count(),
        'this_month': all_cre.filter(created_at__date__gte=month_start).count(),
    }

    # Pagination
    page_size = int(request.query_params.get('page_size', 50))
    page_number = int(request.query_params.get('page', 1))
    paginator = Paginator(qs, page_size)
    page = paginator.get_page(page_number)

    results = []
    for item in page.object_list:
        results.append({
            'id': item.id,
            'name': item.name,
            'phone_number': item.phone_number,
            'location': item.location,
            'notes': item.notes,
            'call_notes': item.call_notes,
            'status': item.status,
            'created_by_name': item.created_by_name,
            'assigned_to_cre_name': item.assigned_to_cre_name,
            'created_at': item.created_at.isoformat(),
        })

    return Response({
        'results': results,
        'count': paginator.count,
        'page': page.number,
        'total_pages': paginator.num_pages,
        'stats': stats,
    })


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def cre_update_lead_status(request, pk):
    """CRE updates lead status and notes."""
    user = request.user
    try:
        if user.role in ('admin', 'team_lead'):
            lead = BOELead.objects.get(pk=pk)
        else:
            lead = BOELead.objects.get(pk=pk, assigned_to_cre=user)
    except BOELead.DoesNotExist:
        return Response({'error': 'Lead not found.'}, status=status.HTTP_404_NOT_FOUND)

    if 'status' in request.data:
        lead.status = request.data['status']
    if 'call_notes' in request.data:
        lead.call_notes = request.data['call_notes']
    if 'notes' in request.data:
        lead.notes = request.data['notes']

    lead.save()
    return Response({'id': lead.id, 'status': lead.status, 'message': 'Lead updated.'})



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def cre_create_lead(request):
    """CRE creates a new lead manually."""
    user = request.user
    name = request.data.get('name', '').strip()
    phone_number = request.data.get('phone_number', '').strip()
    location = request.data.get('location', '').strip()
    notes = request.data.get('notes', '').strip()

    if not name or not phone_number:
        return Response({'error': 'Name and phone number are required.'}, status=status.HTTP_400_BAD_REQUEST)

    lead = BOELead.objects.create(
        name=name,
        phone_number=phone_number,
        location=location,
        notes=notes,
        status='assigned_cre',
        assigned_to_cre=user,
        created_by=user,
        company=user.company,
    )

    return Response({
        'id': lead.id,
        'name': lead.name,
        'message': 'Lead created successfully.',
    }, status=status.HTTP_201_CREATED)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def cre_edit_lead(request, pk):
    """CRE edits a lead assigned to them."""
    user = request.user
    try:
        if user.role in ('admin', 'team_lead'):
            lead = BOELead.objects.get(pk=pk)
        else:
            lead = BOELead.objects.get(pk=pk, assigned_to_cre=user)
    except BOELead.DoesNotExist:
        return Response({'error': 'Lead not found.'}, status=status.HTTP_404_NOT_FOUND)

    if 'name' in request.data:
        lead.name = request.data['name']
    if 'phone_number' in request.data:
        lead.phone_number = request.data['phone_number']
    if 'location' in request.data:
        lead.location = request.data['location']
    if 'notes' in request.data:
        lead.notes = request.data['notes']
    if 'call_notes' in request.data:
        lead.call_notes = request.data['call_notes']
    if 'status' in request.data:
        lead.status = request.data['status']

    lead.save()
    return Response({'id': lead.id, 'message': 'Lead updated.'})


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def cre_delete_lead(request, pk):
    """CRE deletes a lead assigned to them."""
    user = request.user
    try:
        if user.role in ('admin', 'team_lead'):
            lead = BOELead.objects.get(pk=pk)
        else:
            lead = BOELead.objects.get(pk=pk, assigned_to_cre=user)
    except BOELead.DoesNotExist:
        return Response({'error': 'Lead not found.'}, status=status.HTTP_404_NOT_FOUND)

    lead.delete()
    return Response({'message': 'Lead deleted.'}, status=status.HTTP_204_NO_CONTENT)



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def cre_convert_to_task(request, pk):
    """Convert a CRE lead into a task (follow-up call/meeting)."""
    from ase_leads.models.task import ASELeadTask

    user = request.user
    try:
        if user.role in ('admin', 'team_lead'):
            lead = BOELead.objects.get(pk=pk)
        else:
            lead = BOELead.objects.get(pk=pk, assigned_to_cre=user)
    except BOELead.DoesNotExist:
        return Response({'error': 'Lead not found.'}, status=status.HTTP_404_NOT_FOUND)

    title = request.data.get('title', f'Follow up with {lead.name}').strip()
    task_type = request.data.get('task_type', 'call').strip()
    priority = request.data.get('priority', 'medium').strip()
    due_date = request.data.get('due_date')
    description = request.data.get('description', f'Phone: {lead.phone_number}\nLocation: {lead.location}\nNotes: {lead.notes}').strip()

    if not due_date:
        # Default to tomorrow if not provided
        from datetime import timedelta
        from django.utils import timezone
        due_date = (timezone.now() + timedelta(days=1)).isoformat()

    # Get the original BRE creator from source_research if available
    original_creator = lead.created_by  # BOE employee who created the BOELead
    boe_assigner = lead.created_by  # BOE employee who assigned to CRE

    # If the lead has source_research, the original creator is the BRE employee
    if lead.source_research and lead.source_research.created_by:
        original_creator = lead.source_research.created_by  # BRE employee

    task = ASELeadTask.objects.create(
        title=title,
        description=description,
        task_type=task_type,
        priority=priority,
        due_date=due_date,
        assigned_to=user,
        created_by=original_creator,  # BRE employee who created the original data
        assigned_by=boe_assigner,  # BOE employee who assigned the lead
        closed_by=None,  # Will be set when task is marked as done
        status='pending',
    )

    # Mark the lead as completed (converted to task)
    lead.status = 'completed'
    lead.save(update_fields=['status'])

    return Response({
        'id': task.id,
        'title': task.title,
        'message': 'Task created from lead successfully.',
    }, status=status.HTTP_201_CREATED)



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def boe_leads_bulk_delete(request):
    """Bulk delete BOE leads. Supports select_all with filters or specific IDs."""
    user = request.user
    if user.role not in ('admin', 'manager', 'team_lead'):
        return Response({'error': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

    select_all = request.data.get('select_all', False)
    ids = request.data.get('ids', [])

    if select_all:
        from django.db.models import Q
        from accounts.models import Company
        if user.role == 'admin':
            ase_company = Company.objects.filter(code__in=['ASE', 'ASE_TECH']).first()
            qs = BOELead.objects.filter(company=ase_company) if ase_company else BOELead.objects.all()
        else:
            qs = BOELead.objects.filter(company=user.company) if user.company else BOELead.objects.all()
        search = request.data.get('search', '').strip()
        if search:
            qs = qs.filter(Q(name__icontains=search) | Q(phone_number__icontains=search) | Q(location__icontains=search))
        status_filter = request.data.get('status', '').strip()
        if status_filter and status_filter != 'all':
            qs = qs.filter(status=status_filter)
        deleted = qs.count()
        qs.delete()
    elif ids:
        deleted = BOELead.objects.filter(id__in=ids).delete()[0]
    else:
        return Response({'error': 'No records selected.'}, status=status.HTTP_400_BAD_REQUEST)

    return Response({'message': f'{deleted} records deleted.', 'deleted': deleted})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def boe_leads_bulk_assign(request):
    """Bulk assign BOE leads to CRE. Supports select_all with filters or specific IDs."""
    user = request.user
    if user.role not in ('admin', 'manager', 'team_lead', 'employee'):
        return Response({'error': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

    from django.contrib.auth import get_user_model
    User = get_user_model()

    assigned_to_cre_id = request.data.get('assigned_to_cre')
    if not assigned_to_cre_id:
        return Response({'error': 'assigned_to_cre is required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        cre_user = User.objects.get(pk=assigned_to_cre_id)
    except User.DoesNotExist:
        return Response({'error': 'CRE user not found.'}, status=status.HTTP_400_BAD_REQUEST)

    select_all = request.data.get('select_all', False)
    ids = request.data.get('ids', [])

    if select_all:
        from django.db.models import Q
        if user.role == 'admin':
            from accounts.models import Company
            ase_company = Company.objects.filter(code__in=['ASE', 'ASE_TECH']).first()
            qs = BOELead.objects.filter(company=ase_company) if ase_company else BOELead.objects.all()
        elif user.role in ('manager', 'team_lead'):
            qs = BOELead.objects.filter(company=user.company) if user.company else BOELead.objects.all()
        else:
            # Employees can only assign their own leads
            qs = BOELead.objects.filter(created_by=user)
        search = request.data.get('search', '').strip()
        if search:
            qs = qs.filter(Q(name__icontains=search) | Q(phone_number__icontains=search) | Q(location__icontains=search))
        status_filter = request.data.get('status', '').strip()
        if status_filter and status_filter != 'all':
            qs = qs.filter(status=status_filter)
        updated = qs.update(assigned_to_cre=cre_user, status='assigned_cre')
    elif ids:
        if user.role in ('admin', 'manager', 'team_lead'):
            updated = BOELead.objects.filter(id__in=ids).update(assigned_to_cre=cre_user, status='assigned_cre')
        else:
            # Employees can only assign their own leads
            updated = BOELead.objects.filter(id__in=ids, created_by=user).update(assigned_to_cre=cre_user, status='assigned_cre')
    else:
        return Response({'error': 'No records selected.'}, status=status.HTTP_400_BAD_REQUEST)

    return Response({'message': f'{updated} records assigned.', 'updated': updated})



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def boe_leads_creators(request):
    """Return distinct users who created or are assigned CRE for BOE leads (for filter dropdown)."""
    from django.contrib.auth import get_user_model
    from django.db.models import Q
    User = get_user_model()

    user = request.user
    if user.role in ('admin', 'manager', 'team_lead'):
        # Get both creators and CRE-assigned users
        creator_ids = set(BOELead.objects.values_list('created_by_id', flat=True).distinct())
        cre_ids = set(BOELead.objects.filter(assigned_to_cre__isnull=False).values_list('assigned_to_cre_id', flat=True).distinct())
        all_ids = creator_ids | cre_ids
    else:
        creator_ids = set(BOELead.objects.filter(created_by=user).values_list('created_by_id', flat=True).distinct())
        all_ids = creator_ids

    users = User.objects.filter(id__in=all_ids, is_active=True).values(
        'id', 'first_name', 'last_name', 'username', 'role'
    )
    result = []
    for u in users:
        name = f"{u['first_name']} {u['last_name']}".strip() or u['username']
        role_label = u['role'].replace('_', ' ').title() if u['role'] not in ('employee',) else ''
        display_name = f"{name} ({role_label})" if role_label else name
        result.append({'id': u['id'], 'first_name': u['first_name'], 'last_name': u['last_name'], 'name': display_name})
    return Response(result)



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bre_research_auto_assign(request):
    """
    Auto-assign unassigned (new) BRE research data equally to BOE employees.
    Distributes in round-robin fashion.
    
    Request body:
      - boe_user_ids (optional): List of specific BOE user IDs to assign to. If empty, assigns to all BOE members.
      - limit (optional): Max number of records to assign. If empty, assigns all unassigned.
    """
    from django.contrib.auth import get_user_model
    from teams.models import Team
    User = get_user_model()

    user = request.user
    company = user.company
    # For admin users, default to ASE company for marketing panel operations
    if user.role == 'admin':
        from accounts.models import Company
        ase_company = Company.objects.filter(code__in=['ASE', 'ASE_TECH']).first()
        if ase_company:
            company = ase_company
    if not company:
        company_id = request.data.get('company')
        if company_id:
            from accounts.models import Company
            try:
                company = Company.objects.get(id=company_id)
            except Company.DoesNotExist:
                return Response({'error': 'Company not found.'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'error': 'Company is required.'}, status=status.HTTP_400_BAD_REQUEST)

    # Get BOE users to assign to
    boe_user_ids = request.data.get('boe_user_ids', [])
    if boe_user_ids:
        boe_users = list(User.objects.filter(id__in=boe_user_ids, is_active=True))
    else:
        # Get all BOE team members
        boe_teams = Team.objects.filter(marketing_category='boe', is_active=True, company=company)
        boe_users = list(User.objects.filter(team__in=boe_teams, is_active=True))

    if not boe_users:
        return Response({'error': 'No BOE employees found to assign to.'}, status=status.HTTP_400_BAD_REQUEST)

    # Get unassigned records
    unassigned = BREResearchData.objects.filter(company=company, status='new').order_by('created_at')

    # Apply limit if provided
    limit = request.data.get('limit')
    if limit:
        limit = int(limit)
        unassigned = unassigned[:limit]

    # Get IDs to assign
    record_ids = list(unassigned.values_list('id', flat=True))
    total = len(record_ids)

    if total == 0:
        return Response({'error': 'No unassigned records to distribute.'}, status=status.HTTP_400_BAD_REQUEST)

    # Round-robin distribution
    num_users = len(boe_users)
    assigned_counts = {u.id: 0 for u in boe_users}

    for i, record_id in enumerate(record_ids):
        target_user = boe_users[i % num_users]
        BREResearchData.objects.filter(id=record_id).update(
            assigned_to=target_user,
            status='assigned'
        )
        assigned_counts[target_user.id] += 1

    # Build response with assignment details
    details = []
    for u in boe_users:
        name = f"{u.first_name} {u.last_name}".strip() or u.username
        details.append(f"{name}: {assigned_counts[u.id]}")

    return Response({
        'message': f'{total} records auto-assigned to {num_users} BOE employees.',
        'total_assigned': total,
        'distribution': details,
    })
